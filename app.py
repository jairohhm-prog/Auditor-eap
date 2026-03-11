import streamlit as st
import openpyxl
from openpyxl.styles import Font
import io

# 1. CONFIGURAÇÃO DA INTERFACE (MENSAGENS INICIAIS)
st.set_page_config(page_title="Auditor de EAP", layout="centered")
st.title("📊 Automatizador e Auditor de EAP")
st.write("Faça o upload da sua planilha Excel. O sistema validará as regras de preenchimento antes de itemizar.")

st.warning("""
**⚠️ REGRAS DE PREENCHIMENTO:**
Para o sistema funcionar, sua planilha deve conter colunas com os nomes **ITEM**, **DESCRIÇÃO** e **UNIDADE** na região do cabeçalho.

**📌 COMO ITEMIZAR:**
Defina apenas a numeração dos **Títulos** na coluna ITEM, indicando o nível hierárquico:
* **1** (Nível 1)
* **1.1** (Nível 2)
* **1.1.1** (Nível 3) e assim por diante...

Deixe a coluna ITEM **vazia** nas linhas de **Serviço** (as que possuem UNIDADE). O programa fará a numeração delas automaticamente, sem gerar duplicidades!
""")

# 2. BOTÃO DE UPLOAD
arquivo_carregado = st.file_uploader("Selecione a planilha (.xlsx)", type=["xlsx"])

if arquivo_carregado is not None:
    st.info("Arquivo recebido! Analisando estrutura...")
    
    wb = openpyxl.load_workbook(arquivo_carregado)
    planilha = wb.active

    # 3. RADAR DE COLUNAS EXATO
    coluna_item = None
    coluna_descricao = None
    coluna_unidade = None
    linha_inicio_dados = None

    for linha in range(1, 21):
        for coluna in range(1, planilha.max_column + 1):
            valor_celula = str(planilha.cell(row=linha, column=coluna).value).strip().upper()
            if valor_celula == "ITEM":
                coluna_item = coluna
                linha_inicio_dados = linha + 1 
            elif valor_celula in ["DESCRIÇÃO", "DESCRICAO"]:
                coluna_descricao = coluna
            elif valor_celula == "UNIDADE":
                coluna_unidade = coluna
                
        if coluna_item is not None and coluna_descricao is not None and coluna_unidade is not None:
            break

    if coluna_item is None or coluna_descricao is None or coluna_unidade is None:
        st.error("❌ ERRO: Não localizei as colunas 'ITEM', 'DESCRIÇÃO' ou 'UNIDADE' no cabeçalho. Verifique sua planilha e tente novamente.")
    else:
        
        # 4. ETAPA DE VALIDAÇÃO CRÍTICA (Pre-flight Check)
        erros_criticos = []
        
        for linha in range(linha_inicio_dados, planilha.max_row + 1):
            celula_descricao = planilha.cell(row=linha, column=coluna_descricao).value
            
            if celula_descricao is None or str(celula_descricao).strip() == "":
                continue
                
            celula_item = planilha.cell(row=linha, column=coluna_item).value
            celula_unidade = planilha.cell(row=linha, column=coluna_unidade).value
            
            tem_item = celula_item is not None and str(celula_item).strip() != ""
            tem_unidade = celula_unidade is not None and str(celula_unidade).strip() != ""
            descricao_texto = str(celula_descricao).strip()
            
            if not tem_item and not tem_unidade:
                erros_criticos.append(f"❌ Linha {linha}: Título sem nível definido (As colunas 'ITEM' e 'UNIDADE' estão vazias) ➡️ [{descricao_texto}]")
            elif tem_item and tem_unidade:
                erros_criticos.append(f"❌ Linha {linha}: Serviço numerado indevidamente (As colunas 'ITEM' e 'UNIDADE' foram preenchidas juntas) ➡️ [{descricao_texto}]")

        # 5. DECISÃO DO SISTEMA
        if len(erros_criticos) > 0:
            st.error("🛑 AUDITORIA FALHOU: O arquivo contém erros de preenchimento estrutural e não pode ser processado.")
            st.write("Por favor, corrija as seguintes linhas na sua planilha e faça o upload novamente:")
            
            for erro in erros_criticos:
                st.warning(erro)
                
        else:
            # 6. EXECUÇÃO DA LÓGICA MATEMÁTICA UNIFICADA
            st.success("✅ Validação estrutural concluída. Iniciando itemização...")
            hierarquia_atual = [] 
            profundidade_ultimo_titulo = 0 # Guarda a profundidade do título pai
            log_correcoes = [] 

            barra_progresso = st.progress(0)
            total_linhas = planilha.max_row - linha_inicio_dados + 1

            for idx, linha in enumerate(range(linha_inicio_dados, planilha.max_row + 1)):
                celula_descricao = planilha.cell(row=linha, column=coluna_descricao).value
                celula_item = planilha.cell(row=linha, column=coluna_item).value
                celula_unidade = planilha.cell(row=linha, column=coluna_unidade).value

                if celula_descricao is None or str(celula_descricao).strip() == "":
                    continue

                tem_item = celula_item is not None and str(celula_item).strip() != ""
                tem_unidade = celula_unidade is not None and str(celula_unidade).strip() != ""
                descricao_texto = str(celula_descricao).strip()

                if tem_item and not tem_unidade:
                    # É UM TÍTULO
                    item_digitado = str(celula_item).strip()
                    partes = item_digitado.split('.')
                    profundidade_desejada = len(partes)
                    
                    # Atualiza a hierarquia unificada
                    if profundidade_desejada > len(hierarquia_atual):
                        while len(hierarquia_atual) < profundidade_desejada:
                            hierarquia_atual.append(1)
                    elif profundidade_desejada == len(hierarquia_atual):
                        hierarquia_atual[-1] += 1
                    else:
                        hierarquia_atual = hierarquia_atual[:profundidade_desejada]
                        hierarquia_atual[-1] += 1
                        
                    item_correto = ".".join(str(x) for x in hierarquia_atual)
                    profundidade_ultimo_titulo = profundidade_desejada # Título define a nova base para os serviços
                    
                    if item_correto != item_digitado:
                        log_correcoes.append((linha, descricao_texto, item_digitado, item_correto))
                        
                    planilha.cell(row=linha, column=coluna_item).value = item_correto
                    
                elif not tem_item and tem_unidade:
                    # É UM SERVIÇO
                    # Serviço é sempre um nível abaixo do último título pai
                    profundidade_desejada = profundidade_ultimo_titulo + 1 if profundidade_ultimo_titulo > 0 else 1
                    
                    # Aplica a MESMA regra de numeração para o serviço não conflitar com títulos
                    if profundidade_desejada > len(hierarquia_atual):
                        while len(hierarquia_atual) < profundidade_desejada:
                            hierarquia_atual.append(1)
                    elif profundidade_desejada == len(hierarquia_atual):
                        hierarquia_atual[-1] += 1
                    else:
                        hierarquia_atual = hierarquia_atual[:profundidade_desejada]
                        hierarquia_atual[-1] += 1
                        
                    item_correto = ".".join(str(x) for x in hierarquia_atual)
                    
                    # Atenção: Serviços não atualizam 'profundidade_ultimo_titulo' porque serviços não podem ser pais de outros itens
                    
                    planilha.cell(row=linha, column=coluna_item).value = item_correto
                
                # Otimização de Memória (Mantida para planilhas gigantes)
                if idx % 500 == 0 or idx == (total_linhas - 1):
                    progresso_atual = min((idx + 1) / total_linhas, 1.0)
                    barra_progresso.progress(progresso_atual)

            # 7. CRIAÇÃO DA ABA DE LOG NO EXCEL
            if "LOG" in wb.sheetnames:
                wb.remove(wb["LOG"])
            aba_log = wb.create_sheet(title="LOG")
            
            if len(log_correcoes) > 0:
                aba_log.append(["Linha", "Descrição do Item", "O que estava digitado", "Como foi corrigido"])
                for celula in aba_log[1]:
                    celula.font = Font(bold=True)
                for erro in log_correcoes:
                    aba_log.append([erro[0], erro[1], erro[2], erro[3]])
                st.warning(f"⚠️ Foram corrigidas {len(log_correcoes)} inconsistências na numeração matemática (Verifique a aba LOG).")
            else:
                aba_log.append(["Nenhuma correção de nível foi necessária na estrutura da EAP."])

            # 8. PREPARAÇÃO PARA DOWNLOAD
            saida_memoria = io.BytesIO()
            wb.save(saida_memoria)
            saida_memoria.seek(0)

            st.download_button(
                label="📥 Baixar Planilha Itemizada e Auditada",
                data=saida_memoria,
                file_name="EAP_Automatizada.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )